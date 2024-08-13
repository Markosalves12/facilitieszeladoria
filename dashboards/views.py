from django.shortcuts import render, redirect
from utils.utils import paginate, formatar_atributos
import openpyxl
from django.http import HttpResponse
from dashboards.forms import FiltroTableForms, FiltroTableManutencaoForms
from utils.utils import block_view
from dashboards.utils import colect_dados, colect_dados_manutencao_equipamentos, colect_dados_manutencao_ferramentas, colect_dados_planejamento
from servico.models import Servicos
from datetime import datetime
from django.conf import settings
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from io import BytesIO
import os

from utils.utils import capturate_paramns
from django.db.models import Q
from datetime import timedelta
from django.utils import timezone
from django.db.models.functions import Now, TruncDate
import plotly.io as pio

from django.db.models import Sum
from dashboards.view_data import (get_total_area_by_category, plot_horizontal_bar_chart,
                                  plot_grouped_bar_chart, get_total_area_and_counts_by_month)


# função que exporta os dados para o formato excel
def exportar_excel(request, login_type, id, datastart, datafim, empresa, unidade, localidades, negocios):
    # Função que bloqueia a view em caso das variáveis de sessão terem sido alteradas no logout
    block = block_view(
        request,
        login_type=login_type,
        id=id
    )

    if block == True:
        return redirect('logout')

    # Crie um novo arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # cabeçalhos da tabela exportada
    headers = [
        'Tipo de empresa',
        'Empresa',

        # cabeçalhos referentes a tabela de serviços
        'ID Agendamento',
        'Tipo de agendamento',
        'Descrição do Serviço',
        'Colaboradores Chamados',
        'Serviços Solicitados',
        'Data de Início',
        'Data de conclusao',
        'Antes',
        'Depois',
        'Status do Serviço',

        # cabeçalhos referentes a tabela de equipamentos
        'Marca do Equipamento',
        'Equipamento catálogo',
        'Equipamento Empresa',
        'Tipo equipamento',
        'Equipamento id',
        'vida util equipamento (meses)',
        'Data de aquisição equipamento',
        'Data de desmobilização equipamento',
        'Matrícula do Equipamento',

        # cabeçalhos referentes a tabela de ferramentas
        'Marca da Ferramenta',
        'Ferramenta catálogo',
        'Ferramenta empresa',
        'Tipo ferramenta',
        'Ferramenta id',
        'vida util ferramenta (meses)',
        'Data de aquisição ferramenta',
        'Data de desmobilização ferramenta',
        'Matrícula da Ferramenta',

        # cabeçalhos referentes a tabela de materiais
        'Material Aplicado',
        'Material Categoria',
        'Forma de consumo',
        'Quantidade',
        'Origem do material',

        # cabeçalhos referentes a tabela de jardins
        'Área atendida',
        'Periodicidade',
        'Área total',
        'Tipo vegetação',
        'Tipo Terreno',
        'Localidade',
        'Negocio',
        'Unidade',

        # cabeçalhos referentes a tabela fato serviços
        'ID de serviço',
        'Tempo na área',
        'Colaborador envolvido',
        'Principal servico',
        'Data e hora de chagada',
        'Data e hora de retorno'
    ]

    # adiciona os cabeçahos no arquivo
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_title

    # Adicione os dados do relatório ao arquivo Excel
    dados = colect_dados(

    )

    # aplica os filtros enviados como parametro na media que existirem
    if datastart != 'None':
        datastart = datetime.strptime(datastart, '%Y-%m-%d')

    if datafim != 'None':
        datafim = datetime.strptime(datafim, '%Y-%m-%d')

    if empresa != 'None':
        dados = dados.filter(
            empresaprestadora=empresa
        )
    else:
        pass

    if unidade != 'None':
        dados = dados.filter(
            unidade=unidade
        )
    else:
        pass

    if localidades != 'None':
        dados = dados.filter(
            localidade=localidades
        )
    else:
        pass

    if negocios != 'None':
        dados = dados.filter(
            tipodeempresa=negocios
        )
    else:
        pass

    if datastart != 'None':
        dados = dados.filter(
            data_de_inicio__gte=datastart  # Usar __gte para maior ou igual
        )
    else:
        pass


    if datafim != 'None':  # Mudar para datafim ao invés de datastart
        dados = dados.filter(
            data_de_inicio__lte=datafim  # Usar __lte para menor ou igual
        )
    else:
        pass


    # escreve as linhas do arquivo excel antes de ser exportado
    for row_num, row in enumerate(dados, start=2):
        row_data = [
            row.tipodeempresa, row.empresaprestadora,

            row.id_servico, row.tipo_agendamento,

            row.descricao_do_servico, row.colaboradores_chamados, row.servicos_solicitados, row.data_de_inicio,

            row.data_de_conclusao, row.antes, row.depois, row.status_servico,

            row.equipamento_marca, row.equipamento_catalogo, row.equipamento_empresa, row.tipo_equipamento, row.equipamento_id, row.vida_util_equipamento, row.data_aquisicao_equipamento, row.data_desmobilizacao_equipamento, row.matricula_equipamento,

            row.ferramenta_marca, row.ferramenta_catalogo, row.ferramenta_empresa, row.tipo_ferramenta, row.ferramenta_id, row.vida_util_ferramenta, row.data_aquisicao_ferramenta, row.data_desmobilizacao_ferramenta, row.matricula_ferramenta,

            row.material_aplicado, row.material_categoria, row.forma_consumo, row.qtd, row.tipo_material,

            row.area_atendida, row.periodicidade_de_retorno,

            row.area_total, row.tipo_vegetacao, row.tipo_terreno, row.localidade, row.tiponegocio, row.unidade,

            row.id_servico, row.tempo_na_area, row.colaborador_envolvido, row.principalservico,
            row.data_hora_chegada, row.data_hora_retorno
        ]
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(
                row=row_num,
                column=col_num
            )
            cell.value = value

    # Defina o nome do arquivo e o tipo de resposta
    # cria uma resposta http com o tipo de conteuso
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # cria um anexo que ´instalado no lado cliente
    response['Content-Disposition'] = 'attachment; filename="relatorio de servicos.xlsx"'

    # Salve o arquivo Excel
    wb.save(response)

    # retorna a resposta
    # linha obrigatória
    # ou o sistema quebra
    return response


# função que visualiza e filtra os dados que desejam ser baixados via '.xlsx'
def relatoriodeservicos(request):
    login_type = request.session['login_type']
    id = request.session['login_id']


    # Função que bloqueia a view em caso das variáveis de sessão terem sido alteradas no logout
    block = block_view(
        request,
        login_type=login_type,
        id=id
    )

    if block == True:
        return redirect('logout')


    empresas = None
    unidades = None
    localidades = None
    negocios = None
    datastart = None
    datafim = None

    # Adicione os dados do relatório ao arquivo Excel
    dados = colect_dados().filter(
        status_servico = "Concluido"
    )[:50]

    # formulário que envia os filtros para serem aplicados
    forms_filtro = FiltroTableForms()

    if request.method == 'POST':
        form = FiltroTableForms(request.POST)
        if form.is_valid():
            # Processar os dados do formulário
            # os dados do formulário não são enviados para o banco de dados
            # são processado como paramtros para serem aplicados como filtro
            empresas = form.cleaned_data['empresa']
            unidades = form.cleaned_data['unidade']
            localidades = form.cleaned_data.get('localidade')
            datastart = form.cleaned_data.get('datastart')
            datafim = form.cleaned_data.get('datafim')

            dados = colect_dados().filter(
                status_servico="Concluido"
            )
            # aplicação dos filtros na medida que for nescessário
            if empresas:
                dados = dados.filter(
                    empresaprestadora=str(empresas).split("|")[0].strip()
                )
                empresas = str(empresas).split("|")[0].strip()

            if unidades:
                dados = dados.filter(
                    unidade=unidades
                )
                unidades = unidades

            if localidades:
                dados = dados.filter(
                    localidade=str(localidades).split("|")[1].strip(),
                )
                localidades = str(localidades).split("|")[1].strip()


            if negocios:
                dados = dados.filter(
                    localidade=str(localidades).split("|")[2].strip(),
                )
                negocios = str(localidades).split("|")[2].strip()


            if datastart:
                dados = dados.filter(
                    data_de_inicio__gte=datastart  # Usar __gte para maior ou igual
                )

            if datafim:  # Mudar para datafim ao invés de datastart
                dados = dados.filter(
                    data_de_inicio__lte=datafim  # Usar __lte para menor ou igual
                )

    elementos_paginados = paginate(
        request=request,
        data_objects=dados,
        per_page=20
    )

    return render(request, 'controle/dashboards/relatorio_de_servicos.html',
                  {'elementos_paginados': elementos_paginados,
                   'forms_filtro': forms_filtro,
                   'empresas': empresas,
                   'unidades': unidades,
                   'localidades': localidades,
                   'negocios':negocios,
                   'datastart': datastart,
                   'datafim': datafim
                   }
    )


# Função que exporta um pdf com os serviços prestados por cada empresa em cada lugar
def exportar_pdf(request, login_type, id, datastart,
                 datafim, empresa, unidade, localidades,
                 negocios, carteira):
    # Função que bloqueia a view em caso das variáveis de sessão terem sido alteradas no logout
    block = block_view(
        request,
        login_type=login_type,
        id=id
    )

    if block:
        return redirect('logout')

    # inseri os dados no pdf
    dados = Servicos.objects.filter(
        status = "Concluido"
    )

    # aplicada os filtros a medida que for nescessário
    if datastart != 'None':
        datastart = datetime.strptime(datastart, '%Y-%m-%d')
        datastart_apoio = datastart.date().strftime('%d/%m/%Y')

    if datafim != 'None':
        datafim = datetime.strptime(datafim, '%Y-%m-%d')
        datafim_apoio = datafim.date().strftime('%d/%m/%Y')


    if empresa != 'None':
        dados = dados.filter(
            colaboradores_escalados__gerente__gestor__empresa__nome=empresa
        )
    else:
        pass


    if unidade != 'None':
        dados = dados.filter(
            area__unidade_jardim__unidadeoriginial__unidade = unidade
        )
    else:
        unidade = ''

    if localidades != 'None':
        dados = dados.filter(
            area__unidade_jardim__nome=localidades
        )
    else:
        localidades = ''


    if negocios != 'None':
        pass
    else:
        negocios = ''

    if carteira != 'None':
        pass
    else:
        carteira = ''


    if datastart != 'None':
        dados = dados.filter(
            data_inicio__gte=datastart  # Usar __gte para maior ou igual
        )
    else:
        datastart_apoio = ''


    if datafim != 'None':  # Mudar para datafim ao invés de datastart
        dados = dados.filter(
            data_inicio__lte=datafim  # Usar __lte para menor ou igual
        )
    else:
        datafim_apoio = ''


    # cria um buffer para inserir os dados no pdf
    buffer = BytesIO()

    # cria um objeto pdf usando o buffer anterior
    p = canvas.Canvas(
        buffer,
        pagesize=letter
    )
    width, height = letter

    # defini a posição inicial do cursor
    x = 50

    # define o caminho da imagem usada no cabeçalho
    # imagem definida nos arquivos estáticos
    header_image_path = os.path.join(settings.MEDIA_ROOT, 'static/images/logo facilities.png')

    # função que cria o cabeçalho propriamente falado
    def draw_header(c):
        # se o endereço da imagem existir
        if os.path.exists(header_image_path):
            # abri a imagem enviada como parametro
            header_image = ImageReader(header_image_path)
            # captura as dimensões da imagem
            header_img_width, header_img_height = header_image.getSize()
            # centraliza a imagem no topo
            x_centered = (width - header_img_width / 3) / 2
            c.drawImage(header_image, x_centered, height - header_img_height / 3 - 20, width=header_img_width / 3, height=header_img_height / 3, mask='auto')
            image_bottom = height - header_img_height / 2 - 20 - header_img_height / 2 - 20
        else:
            image_bottom = height  # Adjust if image is not found

        # Adiciona o titulo do relatório abaixo da imagem

        # configura da fonte e tamanho do titulo
        # c.setFont("Helvetica-Bold", 16)
        c.setFont("Helvetica-Bold", 14)
        # escrever o titulo do relatório
        c.drawString((width - c.stringWidth(f"Relatório de Serviços {unidade} {localidades}",
                                            "Helvetica-Bold", fontSize=12)) / 2,
                     image_bottom + 20, f"Relatório de Serviços {unidade}/{localidades}")

        c.drawString((width - c.stringWidth(f"{datastart_apoio} - {datafim_apoio}",
                                            "Helvetica-Bold", fontSize=12)) / 2,
                     image_bottom + 0,
                     f"{datastart_apoio} - {datafim_apoio}")

        c.drawString((width - c.stringWidth(f"{carteira}/{negocios}",
                                            "Helvetica-Bold", fontSize=12)) / 2,
                     image_bottom - 20,
                     f"{carteira}  -- {negocios}")

        c.setFont("Helvetica", 12)  # Set font back to normal for the rest of the content

    def draw_footer(c, page_number, is_last_page=False):
        c.setFont("Helvetica", 13)
        if is_last_page:
            last_page_text = "Facilities Suzano - Regional norte"
            c.drawString((width - c.stringWidth(last_page_text, fontSize=12)) / 2, 50, last_page_text)

    # Draw the header for the first page
    draw_header(p)
    y = height - 150  # Adjust starting position for content after the header
    page_number = 1
    p.setFont("Helvetica", 10)

    for dado in dados:
        # Add the data_inicio
        p.setFont('Helvetica-Bold', 10)
        p.drawString(x, y, f"Descrição: {dado.descricao_servico}")
        y -= 20

        p.setFont("Helvetica", 10)
        p.drawString(x, y, f"Data de Início: {dado.data_inicio.strftime('%d/%m/%Y')}",)

        y -= 20

        p.drawString(x, y, f"Data de conclusão: {dado.data_conclusao.strftime('%d/%m/%Y')}")
        y -= 20

        p.drawString(x, y, f"área atendida: {dado.area}")
        y -= 20

        p.drawString(x, y, f"Tamanho da área atendida: {dado.area.area} M²")
        y -= 20

        # Add the servicos_escalados
        p.drawString(x, y, "Serviços Escalados:")
        y -= 20

        servicos = formatar_atributos(
            queryset=dado.servicos_escalados.all(),
            atributo='servico'
        )
        p.drawString(x + 20, y, f"- {servicos}")  # Ajuste conforme o campo do modelo Servicos
        y -= 20

        # Add the colaboradores_escalados
        p.drawString(x, y, "Colaboradores Escalados:")
        y -= 20

        colaborador = formatar_atributos(
            queryset=dado.colaboradores_escalados.all(),
            atributo='nome'
        )
        p.drawString(x + 20, y, f"- {colaborador}")  # Ajuste conforme o campo do modelo Colaboradores
        y -= 20


        def add_images_to_canvas(p, dado, x, y):
            def calculate_new_dimensions(img_width, img_height):
                new_width = img_width / 2.4
                new_height = img_height / 2.4
                return new_width, new_height

            def draw_image(image_path, x, y):
                try:
                    img_data = ImageReader(image_path)
                    img_width, img_height = img_data.getSize()
                    new_width, new_height = calculate_new_dimensions(img_width, img_height)
                    p.drawImage(img_data, x, y - new_height, width=new_width, height=new_height, mask='auto')
                    return new_height
                except Exception as e:
                    p.drawString(x, y, f"Erro ao carregar a imagem: {str(e)}")
                    return 0

            # Draw the first image (foto_inicio)
            if dado.foto_inicio:
                y -= 7
                p.drawString(x, y, "Na solicitção")
                y -= 7
                image_path = os.path.join(settings.MEDIA_ROOT, dado.foto_inicio.name)
            else:
                y -= 7
                p.drawString(x, y, "Na solicitção")
                y -= 7
                image_path = os.path.join(settings.MEDIA_ROOT, 'static/assets/image not found (1).png')

            height1 = draw_image(image_path, x, y)

            # Update y position for the next image
            y -= height1 + 10  # 10 is the space between images

            # Draw the second image (foto)
            if dado.foto:
                y -= 7
                p.drawString(x, y, "Na entrega")
                y -= 7
                image_path = os.path.join(settings.MEDIA_ROOT, dado.foto.name)
            else:
                y -= 7
                p.drawString(x, y, "Na entrega")
                y -= 7
                image_path = os.path.join(settings.MEDIA_ROOT, 'static/assets/image not found (1).png')

            draw_image(image_path, x, y)


        add_images_to_canvas(p, dado, x, y)


        # # Add the foto
        # if dado.foto_inicio:
        #     try:
        #         image_path = os.path.join(settings.MEDIA_ROOT, dado.foto_inicio.name)
        #         img_data = ImageReader(image_path)
        #         img_width, img_height = img_data.getSize()
        #         # Reduzindo pela metade
        #         new_width = img_width / 3
        #         new_height = img_height / 3
        #         aspect_ratio = new_height / float(new_width)
        #         p.drawImage(img_data, x, y - (new_width * aspect_ratio), width=new_width,
        #                     height=new_height, mask='auto')
        #         y -= (new_width * aspect_ratio + 20)
        #     except Exception as e:
        #         p.drawString(x, y, f"Erro ao carregar a imagem: {str(e)}")
        #         y -= 20
        #
        # # Reduzindo a segunda imagem (dado.foto)
        # if dado.foto:
        #     try:
        #         image_path = os.path.join(settings.MEDIA_ROOT, dado.foto.name)
        #         img_data = ImageReader(image_path)
        #         img_width, img_height = img_data.getSize()
        #         # Reduzindo pela metade
        #         new_width = img_width / 3
        #         new_height = img_height / 3
        #         aspect_ratio = new_height / float(new_width)
        #         p.drawImage(img_data, x, y - (new_width * aspect_ratio), width=new_width,
        #                     height=new_height, mask='auto')
        #         y -= (new_width * aspect_ratio + 20)
        #     except Exception as e:
        #         p.drawString(x, y, f"Erro ao carregar a imagem: {str(e)}")
        #         y -= 20


        # Draw the footer on the current page
        draw_footer(p, page_number)

        # Show the current page and prepare for the next record
        p.showPage()
        page_number += 1

        p.setFont("Helvetica", 10)  # Reset font size to 12 for new page content
        y = height - 70


    draw_footer(
        p,
        page_number,
        is_last_page=True
    )

    # Close the PDF object cleanly, and we're done.
    p.showPage()
    p.save()

    # Get the value of the BytesIO buffer and write it to the response.
    buffer.seek(0)

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio de servicos.pdf"'

    return response


def relatoriopdf(request):
    login_type = request.session['login_type']
    id = request.session['login_id']

    block = block_view(
        request,
        login_type=login_type,
        id=id
    )
    if block == True:
        return redirect('logout')

    empresas=None
    unidades = None
    localidades = None
    negocios = None
    carteira = None
    datastart = None
    datafim = None

    dados = Servicos.objects.filter(
        status = "Concluido"
    )[:50]

    forms_filtro = FiltroTableForms()


    if request.method == 'POST':
        form = FiltroTableForms(request.POST)
        if form.is_valid():
            # Processar os dados do formulário
            empresas=form.cleaned_data['empresa']
            unidades = form.cleaned_data['unidade']
            localidades = form.cleaned_data.get('localidade')
            datastart = form.cleaned_data.get('datastart')
            datafim = form.cleaned_data.get('datafim')

            dados = Servicos.objects.filter(
                status="Concluido"
            )

            # if unidades:
            #     dados = dados.filter(
            #         area__unidade_jardim__unidadeoriginial=unidades
            #     )
            #     unidades = unidades

            if empresas:
                dados = dados.filter(
                    colaboradores_escalados__gerente__gestor__empresa__nome=str(empresas).split("|")[0].strip()
                )
                carteira = str(empresas).split("|")[2].strip()
                empresas = str(empresas).split("|")[0].strip()

            if unidades:
                dados = dados.filter(
                    area__unidade_jardim__unidadeoriginial=unidades
                )
                unidades = unidades

            if localidades:
                dados = dados.filter(
                    area__unidade_jardim__nome=str(localidades).split("|")[1].strip()
                )
                negocios = str(localidades).split("|")[2].strip()
                localidades = str(localidades).split("|")[1].strip()

            if datastart:
                dados = dados.filter(
                    data_inicio__gte=datastart  # Usar __gte para maior ou igual
                )

            if datafim:  # Mudar para datafim ao invés de datastart
                dados = dados.filter(
                    data_inicio__lte=datafim  # Usar __lte para menor ou igual
                )

    elementos_paginados = paginate(
        request=request,
        data_objects=dados,
        per_page=5
    )

    return render(request, 'controle/dashboards/relatorio_pdf.html',
                  {'elementos_paginados': elementos_paginados,
                   'forms_filtro': forms_filtro,
                   'empresas': empresas,
                   'unidades': unidades,
                   'carteira': carteira,
                   'localidades': localidades,
                   'negocios': negocios,
                   'datastart': datastart,
                   'datafim': datafim})


def exportar_relatorio_manutencao_equipamentos(request, login_type, id,
                                               datastart, datafim, unidades,
                                               empresas, tipomanutencao):
    block = block_view(
        request,
        login_type=login_type,
        id=id
    )

    if block == True:
        return redirect('logout')

    # Crie um novo arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = [
        'ID equipamento',
        'Data/hora de inicio',
        'Data/hora de retorno',
        'Equipamento catálogo',
        'Matricula de equipamento',
        'Marca',
        'Tipo Manutenção',
        'Tempo em manutenção',
        'Tipo Equipamento',
        'Unidade',
        'Empresa'
    ]


    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(
            row=1,
            column=col_num
        )

        cell.value = header_title

    dados = colect_dados_manutencao_equipamentos()


    if datastart != 'None':
        datastart = datetime.fromisoformat(datastart)

    if datafim != 'None':
        datafim = datetime.fromisoformat(datafim)

    if unidades != 'None':
        dados = dados.filter(
            unidade=unidades
        )
    else:
        pass

    if empresas != 'None':
        dados = dados.filter(
            empresa=empresas
        )
    else:
        pass

    if tipomanutencao != 'None':
        dados = dados.filter(
            tipomanutencao=tipomanutencao
        )
    else:
        pass


    if datastart != 'None':
        dados = dados.filter(
            data_inicio__gte=datastart
        )

    if datafim != 'None':
        dados = dados.filter(
            data_fim__lte=datafim
        )



    for row_num, row in enumerate(dados, start=2):
        row_data = [
            row.id_equipamento, row.data_inicio, row.data_fim, row.equipamento_,
            row.matricula_equipamento, row.equipamento_marca, row.tipomanutencao,
            row.tempo_em_manutencao, row.tipo_equipameanto, row.unidade,
            row.empresa
        ]
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(
                row=row_num,
                column=col_num
            )

            cell.value = value


    # Defina o nome do arquivo e o tipo de resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename="relatorio manutencao equipamentos.xlsx"'

    # Salve o arquivo Excel
    wb.save(response)

    return response


def relatoriomanutencaoequipamentos(request):
    login_type = request.session['login_type']
    id = request.session['login_id']

    block = block_view(
        request,
        login_type=login_type,
        id=id
    )

    if block == True:
        return redirect('logout')

    dados = colect_dados_manutencao_equipamentos(

    )[:50]

    # datastart = datetime.strptime('2024-12-31 23:59', '%Y-%m-%d %H:%M')
    # datafim = datetime.strptime('2024-01-01 00:00', '%Y-%m-%d %H:%M')
    datastart = None
    datafim = None
    unidades = None
    empresas = None
    tipomanutencao = None


    forms_filtro = FiltroTableManutencaoForms()


    if request.method == 'POST':
        form = FiltroTableManutencaoForms(request.POST)
        if form.is_valid():
            pass
            # Processar os dados do formulário
            datastart = form.cleaned_data.get('datastart')
            datafim = form.cleaned_data.get('datafim')
            unidades = form.cleaned_data['unidade']
            empresas = form.cleaned_data.get('empresa')
            tipomanutencao = form.cleaned_data.get('tipo_manutencao')

            dados = colect_dados_manutencao_equipamentos()

            if unidades:
                dados = dados.filter(
                    unidade=unidades
                )
                unidades = unidades


            if empresas:
                dados = dados.filter(
                    empresa=str(empresas).split("|")[0].strip()
                )
                empresas = str(empresas).split("|")[0].strip()

            if tipomanutencao:
                dados = dados.filter(
                    tipomanutencao = tipomanutencao
                )


            if datastart:
                dados = dados.filter(
                    data_inicio__gte = datastart
                )

            if datafim:
                dados = dados.filter(
                    data_fim__lte = datafim
                )

    elementos_paginados = paginate(
        request=request,
        data_objects=dados,
        per_page=20
    )

    return render(request, 'controle/dashboards/relatorio_manutencao_equipamentos.html',
                  {'elementos_paginados': elementos_paginados,
                   'forms_filtro': forms_filtro,
                   'datastart': datastart,
                   'datafim': datafim,
                   'unidades': unidades,
                   'empresas': empresas,
                   'tipomanutencao':tipomanutencao})


def exportar_relatorio_manutencao_ferramentas(request, login_type, id,
                                               datastart, datafim, unidades,
                                               empresas, tipomanutencao):
    block = block_view(
        request,
        login_type=login_type,
        id=id
    )

    if block == True:
        return redirect('logout')

    # Crie um novo arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = [
        'ID ferramenta',
        'Data/hora de inicio',
        'Data/hora de retorno',
        'Ferramenta catálogo',
        'Matricula da ferramenta',
        'Marca',
        'Tipo Manutenção',
        'Tempo em manutenção',
        'Tipo Ferramenta',
        'Unidade',
        'Empresa'
    ]


    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(
            row=1,
            column=col_num
        )

        cell.value = header_title


    dados = colect_dados_manutencao_ferramentas()


    if datastart != 'None':
        datastart = datetime.fromisoformat(datastart)

    if datafim != 'None':
        datafim = datetime.fromisoformat(datafim)

    if unidades != 'None':
        dados = dados.filter(
            unidade=unidades
        )
    else:
        pass

    if empresas != 'None':
        dados = dados.filter(
            empresa=empresas
        )
    else:
        pass

    if tipomanutencao != 'None':
        dados = dados.filter(
            tipomanutencao=tipomanutencao
        )
    else:
        pass


    if datastart != 'None':
        dados = dados.filter(
            data_inicio__gte=datastart
        )

    if datafim != 'None':
        dados = dados.filter(
            data_fim__lte=datafim
        )



    for row_num, row in enumerate(dados, start=2):
        row_data = [
            row.id_ferramenta, row.data_inicio, row.data_fim, row.ferramenta_,
            row.matricula_ferramenta, row.ferramenta_marca, row.tipomanutencao,
            row.tempo_em_manutencao, row.tipo_ferramenta, row.unidade,
            row.empresa
        ]
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(
                row=row_num,
                column=col_num
            )
            cell.value = value


    # Defina o nome do arquivo e o tipo de resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="relatorio manutencao ferramentas.xlsx"'

    # Salve o arquivo Excel
    wb.save(response)

    return response


def relatoriomanutencaoferramentas(request):
    login_type = request.session['login_type']
    id = request.session['login_id']

    block = block_view(
        request,
        login_type=login_type,
        id=id
    )

    if block == True:
        return redirect('logout')

    dados = colect_dados_manutencao_ferramentas()[:50]

    datastart = None
    datafim = None
    tipomanutencao = None
    empresas = None
    unidades = None

    forms_filtro = FiltroTableManutencaoForms()


    if request.method == 'POST':
        form = FiltroTableManutencaoForms(request.POST)
        if form.is_valid():
            pass
            # Processar os dados do formulário
            datastart = form.cleaned_data.get('datastart')
            datafim = form.cleaned_data.get('datafim')
            unidades = form.cleaned_data['unidade']
            empresas = form.cleaned_data.get('empresa')
            tipomanutencao = form.cleaned_data.get('tipo_manutencao')
            dados = colect_dados_manutencao_ferramentas()
            if unidades:
                dados = dados.filter(
                    unidade=unidades
                )
                unidades = unidades


            if empresas:
                dados = dados.filter(
                    empresa=str(empresas).split("|")[0].strip()
                )
                empresas = str(empresas).split("|")[0].strip()

            if tipomanutencao:
                dados = dados.filter(
                    tipomanutencao = tipomanutencao
                )


            if datastart:
                dados = dados.filter(
                    data_inicio__gte = datastart
                )

            if datafim:
                dados = dados.filter(
                    data_fim__lte = datafim
                )


    elementos_paginados = paginate(
        request=request,
        data_objects=dados,
        per_page=20
    )

    return render(request, 'controle/dashboards/relatorio_manutencao_ferramentas.html',
                  {'elementos_paginados': elementos_paginados,
                   'forms_filtro': forms_filtro,
                   'datastart': datastart,
                   'datafim': datafim,
                   'unidades': unidades,
                   'empresas': empresas,
                   'tipomanutencao': tipomanutencao}
                  )



def get_total_area(queryset, area_field='area__area'):
    aggregate_result = queryset.aggregate(total_area=Sum(area_field))
    return aggregate_result['total_area'] if aggregate_result['total_area'] else 0


def calculate_areas_and_counts(dados_servicos, status, date_filter=None):
    """
    Calcula a soma das áreas e as contagens de itens para um conjunto de dados filtrados.
    """
    queryset = dados_servicos.filter(status=status)
    if date_filter:
        queryset = queryset.filter(date_filter)
    return get_total_area(queryset)


def generate_chart(dados_servicos, status, field_name, title, label_type, color):
    """
    Gera um gráfico de barras horizontais para um conjunto de dados filtrados.
    """
    area, counts = get_total_area_by_category(dados_servicos.filter(status=status), field_name)
    return plot_horizontal_bar_chart(area, title, 'Área Total', label_type, counts, color)


def generate_grouped_chart(dados_servicos, status, field_name, title, label_type):
    """
    Gera um gráfico de barras agrupadas para um conjunto de dados filtrados.
    """
    area, counts, categories = get_total_area_and_counts_by_month(dados_servicos.filter(status=status), field_name)
    return plot_grouped_bar_chart(area, title, 'Mês', 'Área Total', counts, categories, label_type)


def dashboard_gerencial(request):
    login_type = request.session['login_type']
    id = request.session['login_id']

    block = block_view(request, login_type=login_type, id=id)
    if block:
        return redirect('logout')

    empresa = capturate_paramns(login_type=login_type, id=id, type='empresa')
    forms_filtro = FiltroTableForms()

    datastart = None
    datafim = None

    if login_type == 'Colaborador':
        filter_query = Q(colaboradores_escalados__id_random=id)
    elif login_type == 'Gerente':
        filter_query = Q(colaboradores_escalados__gerente__id_random=id)
    elif login_type == 'Gestor':
        filter_query = Q(colaboradores_escalados__gerente__gestor__id_random=id)
    else:
        filter_query = Q()

    dados_servicos = (
        Servicos.objects.filter(status__in=['Agendado', 'Em andamento'])
        .annotate(data_atual=TruncDate(Now()))
        .filter(filter_query)
        .order_by('-data_inicio')
        .distinct()
    )

    if request.method == 'POST':
        form = FiltroTableForms(request.POST)
        if form.is_valid():
            # Processar os dados do formulário
            # os dados do formulário não são enviados para o banco de dados
            # são processado como paramtros para serem aplicados como filtro
            datastart = form.cleaned_data.get('datastart')
            datafim = form.cleaned_data.get('datafim')

            if datastart:
                dados_servicos = dados_servicos.filter(
                    data_inicio__gte = datastart
                )

            if datafim:
                dados_servicos = dados_servicos.filter(
                    data_inicio__lte = datafim
                )

    em_andamento = dados_servicos.filter(status='Em andamento').count()
    atrasados = dados_servicos.filter(data_inicio__lt=timezone.now().date()).exclude(status="Em andamento").count()

    one_day = timezone.now().date() + timedelta(days=1)
    seven_days = timezone.now().date() + timedelta(days=7)

    proximos = dados_servicos.filter(
        data_inicio__gte=one_day,
        data_inicio__lte=seven_days
    ).exclude(
        status="Em andamento",
        data_inicio__lte=seven_days
    ).count()

    agendamentos = dados_servicos.filter(
        status='Agendado',
        data_inicio__gte=seven_days,
    ).count()

    # elementos_paginados = paginate(request=request, data_objects=dados_servicos, per_page=10)

    total_de_areas_agendadas = calculate_areas_and_counts(dados_servicos, 'Agendado', Q(data_inicio__gte=seven_days))
    total_de_areas_atrasadas = calculate_areas_and_counts(dados_servicos, 'Agendado', Q(data_inicio__lt=timezone.now().date()))
    total_de_areas_proximas = calculate_areas_and_counts(dados_servicos, 'Agendado', Q(data_inicio__gte=one_day) & Q(data_inicio__lte=seven_days))
    total_de_areas_em_andamento = calculate_areas_and_counts(dados_servicos, 'Em andamento')

    fig_charts = {
        'fig_terreno': generate_chart(dados_servicos.filter(data_inicio__gte=seven_days,), 'Agendado', 'area__terreno', 'Área Total por Tipo de Terreno (Agendado)', 'terreno', '#badbff').to_html(full_html=False),
        'fig_vegetacao': generate_chart(dados_servicos.filter(data_inicio__gte=seven_days,), 'Agendado', 'area__vegetacao', 'Área Total por Tipo de Vegetação (Agendado)', 'vegetação', '#90ee90').to_html(full_html=False),
        'fig_localidade': generate_chart(dados_servicos.filter(data_inicio__gte=seven_days,), 'Agendado', 'area__unidade_jardim__nome', 'Área Total por localidade (Agendado)', 'localidade', '#90ee90').to_html(full_html=False),
        'fig_terreno_andamento': generate_chart(dados_servicos.filter(status='Em andamento'), 'Em andamento', 'area__terreno', 'Área Total por Tipo de Terreno (Em andamento)', 'terreno', '#badbff').to_html(full_html=False),
        'fig_vegetacao_em_andamento': generate_chart(dados_servicos.filter(status='Em andamento'), 'Em andamento', 'area__vegetacao', 'Área Total por Tipo de Vegetação (Em andamento)', 'vegetação', '#90ee90').to_html(full_html=False),
        'fig_terreno_proximo': generate_chart(
            dados_servicos.filter(
                data_inicio__gte=one_day,
                data_inicio__lte=seven_days
            ).exclude(
                status="Em andamento",
                data_inicio__lte=seven_days
            ), 'Agendado', 'area__terreno', 'Área Total por Tipo de Terreno (Próximo)', 'terreno', '#badbff'
        ).to_html(full_html=False),
        'fig_vegetacao_proximo': generate_chart(dados_servicos.filter(
                data_inicio__gte=one_day,
                data_inicio__lte=seven_days
            ).exclude(
                status="Em andamento",
                data_inicio__lte=seven_days
            ), 'Agendado', 'area__vegetacao', 'Área Total por Tipo de Vegetação (Próximo)', 'vegetação', '#90ee90'
        ).to_html(full_html=False),
    }

    fig_grouped_charts = {
        'fig_mes_html': generate_grouped_chart(dados_servicos.filter(data_inicio__gte=seven_days,), 'Agendado', 'area__vegetacao', 'Serviços por Mês/vegetação (Agendado)', 'vegetação').to_html(full_html=False),
        'fig_mes_html_terreno': generate_grouped_chart(dados_servicos.filter(data_inicio__gte=seven_days,), 'Agendado', 'area__terreno', 'Serviços por Mês/terreno (Agendado)', 'terreno').to_html(full_html=False),
        'fig_mes_html_servico': generate_grouped_chart(dados_servicos.filter(data_inicio__gte=seven_days,), 'Agendado', 'servicos_escalados__servico', 'Serviços por Mês/serviço (Agendado)', 'serviço').to_html(full_html=False),
    }

    return render(request, 'controle/dashboards/gerencial.html', {
        # 'elementos_paginados': elementos_paginados,
        'agendamentos': agendamentos,
        'em_andamento': em_andamento,
        'atrasados': atrasados,
        'proximos': proximos,
        'empresa': empresa,
        'total_de_areas_agendadas': total_de_areas_agendadas,
        'total_de_areas_atrasadas': total_de_areas_atrasadas,
        'total_de_areas_proximas': total_de_areas_proximas,
        'total_de_areas_em_andamento': total_de_areas_em_andamento,
        'forms_filtro': forms_filtro,
        'datastart': datastart,
        'datafim': datafim,
        **fig_charts,
        **fig_grouped_charts,
    })

def save_plotly_fig_as_image(fig, file_path):
    """Save Plotly figure as an image."""
    pio.write_image(fig, file_path)


def exportar_relatorio_de_planejamento(request, login_type, id, datastart, datafim):
    login_type = request.session['login_type']
    id = request.session['login_id']

    block = block_view(request, login_type=login_type, id=id)
    if block:
        return redirect('logout')

    if login_type == 'Colaborador':
        filter_query = Q(colaboradores_escalados__id_random=id)
    elif login_type == 'Gerente':
        filter_query = Q(colaboradores_escalados__gerente__id_random=id)
    elif login_type == 'Gestor':
        filter_query = Q(colaboradores_escalados__gerente__gestor__id_random=id)
    else:
        filter_query = Q()

    dados_servicos = (
        Servicos.objects.filter(status__in=['Agendado', 'Em andamento'])
        .annotate(data_atual=TruncDate(Now()))
        .filter(filter_query)
        .order_by('-data_inicio')
        .distinct()
    )

    one_day = timezone.now().date() + timedelta(days=1)
    seven_days = timezone.now().date() + timedelta(days=7)

    # aplicada os filtros a medida que for nescessário
    if datastart != 'None':
        datastart = datetime.strptime(datastart, '%Y-%m-%d')
        datastart_apoio = datastart.date().strftime('%d/%m/%Y')

    if datafim != 'None':
        datafim = datetime.strptime(datafim, '%Y-%m-%d')
        datafim_apoio = datafim.date().strftime('%d/%m/%Y')

    if datastart != 'None':
        dados_servicos = dados_servicos.filter(
            data_inicio__gte=datastart  # Usar __gte para maior ou igual
        )
    else:
        datastart_apoio = ''


    if datafim != 'None':  # Mudar para datafim ao invés de datastart
        dados_servicos = dados_servicos.filter(
            data_inicio__lte=datafim  # Usar __lte para menor ou igual
        )
    else:
        datafim_apoio = ''

    # cria um buffer para inserir os dados no pdf
    buffer = BytesIO()

    # cria um objeto pdf usando o buffer anterior
    p = canvas.Canvas(
        buffer,
        pagesize=letter
    )
    width, height = letter

    # Variáveis para ajustes
    x_start = 50  # Posição inicial x
    line_height = 20  # Altura das linhas de texto
    max_items_per_page = 2  # Máximo de objetos 'dados_servicos' por página

    # Cria um buffer para inserir os dados no PDF
    buffer = BytesIO()

    # Cria um objeto PDF usando o buffer anterior
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Define o caminho da imagem usada no cabeçalho (ajuste conforme necessário)
    header_image_path = os.path.join(settings.MEDIA_ROOT, 'static/images/logo facilities.png')

    # Função que cria o cabeçalho propriamente falado
    def draw_header(c):
        if os.path.exists(header_image_path):
            header_image = ImageReader(header_image_path)
            header_img_width, header_img_height = header_image.getSize()
            x_centered = (width - header_img_width / 3) / 2
            c.drawImage(header_image, x_centered, height - header_img_height / 3 - 20,
                        width=header_img_width / 3, height=header_img_height / 3, mask='auto')
            image_bottom = height - header_img_height / 2 - 20 - header_img_height / 2 - 20
        else:
            image_bottom = height

        c.setFont("Helvetica-Bold", 14)
        c.drawString(
            (width - c.stringWidth("Relatório de planejamento de serviços", "Helvetica-Bold", fontSize=12)) / 2,
            image_bottom + 20, "Relatório de planejamento de serviços")

        c.drawString((width - c.stringWidth(f"{datastart_apoio} - {datafim_apoio}", "Helvetica-Bold", fontSize=12)) / 2,
                     image_bottom + 0, f"{datastart_apoio} - {datafim_apoio}")

        c.setFont("Helvetica", 12)

    # Função para desenhar o rodapé
    def draw_footer(c, page_number, is_last_page=False):
        c.setFont("Helvetica", 13)
        if is_last_page:
            last_page_text = "Facilities Suzano - Regional norte"
            c.drawString((width - c.stringWidth(last_page_text, fontSize=12)) / 2, 50, last_page_text)

    # Função para desenhar os dados de um serviço
    def draw_service_data(c, dado, y):
        c.setFont('Helvetica-Bold', 10)
        c.drawString(x_start, y, f"Descrição: {dado.descricao_servico}")
        y -= line_height

        c.setFont("Helvetica", 10)
        c.drawString(x_start, y, f"Data de Início: {dado.data_inicio.strftime('%d/%m/%Y')}")
        y -= line_height

        c.drawString(x_start, y, f"Área atendida: {dado.area}")
        y -= line_height

        c.drawString(x_start, y, f"Tamanho da área atendida: {dado.area.area} M²")
        y -= line_height

        c.drawString(x_start, y, "Serviços Escalados:")
        y -= line_height

        servicos = formatar_atributos(dado.servicos_escalados.all(), 'servico')
        c.drawString(x_start + 20, y, f"- {servicos}")
        y -= line_height

        c.drawString(x_start, y, "Colaboradores Escalados:")
        y -= line_height

        colaborador = formatar_atributos(dado.colaboradores_escalados.all(), 'nome')
        c.drawString(x_start + 20, y, f"- {colaborador}")
        y -= 30  # Espaçamento entre os itens

        return y

    # Função para desenhar a lista de serviços com paginação
    def draw_service_list(c, data_list, start_y, start_page):
        y = start_y
        page_number = start_page
        for index, dado in enumerate(data_list):
            if index > 0 and index % max_items_per_page == 0:
                draw_footer(c, page_number, is_last_page=False)
                c.showPage()
                page_number += 1
                draw_header(c)
                y = height - 150  # Reinicia a posição y para o novo cabeçalho

            y = draw_service_data(c, dado, y)

        return y, page_number

    def add_figures_to_pdf(c, fig_dict, start_y, start_page):
        width, height = letter
        y = start_y
        page_number = start_page

        # Define o caminho do diretório onde as imagens serão salvas usando BASE_DIR
        image_dir = os.path.join(settings.BASE_DIR, 'media', 'images')

        # Cria o diretório se ele não existir
        if not os.path.exists(image_dir):
            os.makedirs(image_dir)

        for fig_name, fig_html in fig_dict.items():
            fig_file_path = os.path.join(image_dir, f'{fig_name}.png')
            save_plotly_fig_as_image(fig_html, fig_file_path)

            if os.path.exists(fig_file_path):
                fig_image = ImageReader(fig_file_path)
                fig_img_width, fig_img_height = fig_image.getSize()
                if y - fig_img_height / 3 < 50:
                    draw_footer(c, page_number, is_last_page=False)
                    c.showPage()
                    page_number += 1
                    draw_header(c)
                    y = height - 150

                n = 2
                c.drawImage(fig_image, (width - fig_img_width / n) / 2, y - fig_img_height / n,
                            width=fig_img_width / n, height=fig_img_height / n, mask='auto')
                y -= fig_img_height / n + 20

        return y, page_number


    fig_charts = {
        'fig_terreno': generate_chart(dados_servicos.filter(data_inicio__gte=seven_days), 'Agendado', 'area__terreno',
                                      'Área Total por Tipo de Terreno (Agendado)', 'terreno', '#badbff'),
        'fig_vegetacao': generate_chart(dados_servicos.filter(data_inicio__gte=seven_days), 'Agendado',
                                        'area__vegetacao', 'Área Total por Tipo de Vegetação (Agendado)', 'vegetação',
                                        '#90ee90'),
        'fig_localidade': generate_chart(dados_servicos.filter(data_inicio__gte=seven_days), 'Agendado',
                                         'area__unidade_jardim__nome', 'Área Total por localidade (Agendado)',
                                         'localidade', '#90ee90'),
    }


    fig_charts_proximos = {
        'fig_terreno': generate_chart(
            dados_servicos.filter(
                data_inicio__gte=one_day,
                data_inicio__lte=seven_days
            ).exclude(
                status="Em andamento",
                data_inicio__lte=seven_days
            ), 'Agendado', 'area__terreno', 'Área Total por Tipo de Terreno (Próximo)', 'terreno', '#badbff'
        ),
        'fig_vegetacao': generate_chart(dados_servicos.filter(
                data_inicio__gte=one_day,
                data_inicio__lte=seven_days
            ).exclude(
                status="Em andamento",
                data_inicio__lte=seven_days
            ), 'Agendado', 'area__vegetacao', 'Área Total por Tipo de Vegetação (Próximo)', 'vegetação', '#90ee90'
        ),
        'fig_localidade': generate_chart(dados_servicos.filter(
                data_inicio__gte=one_day,
                data_inicio__lte=seven_days
            ).exclude(
                status="Em andamento",
                data_inicio__lte=seven_days
            ), 'Agendado', 'area__unidade_jardim__nome', 'Área Total por localidade (Próximo)', 'localidade', '#90ee90'
        ),
    }

    fig_charts_endamento = {
    'fig_terreno_andamento': generate_chart(dados_servicos.filter(status='Em andamento'), 'Em andamento',
                                            'area__terreno', 'Área Total por Tipo de Terreno (Em andamento)',
                                            'terreno', '#badbff'),
    'fig_vegetacao_em_andamento': generate_chart(dados_servicos.filter(status='Em andamento'), 'Em andamento',
                                                 'area__vegetacao',
                                                 'Área Total por Tipo de Vegetação (Em andamento)', 'vegetação',
                                                 '#90ee90')
    }

    fig_grouped_charts = {
        'fig_mes_html': generate_grouped_chart(dados_servicos.filter(data_inicio__gte=seven_days), 'Agendado',
                                               'area__vegetacao', 'Serviços por Mês/vegetação (Agendado)', 'vegetação'),
        'fig_mes_html_terreno': generate_grouped_chart(dados_servicos.filter(data_inicio__gte=seven_days), 'Agendado',
                                                       'area__terreno', 'Serviços por Mês/terreno (Agendado)',
                                                       'terreno'),
        'fig_mes_html_servico': generate_grouped_chart(dados_servicos.filter(data_inicio__gte=seven_days), 'Agendado',
                                                       'servicos_escalados__servico',
                                                       'Serviços por Mês/serviço (Agendado)', 'serviço'),
    }

    # Desenha o cabeçalho inicial
    draw_header(p)
    start_y = height - 150  # Posição inicial para o conteúdo após o cabeçalho
    start_page = 1

    p.setFont('Helvetica-Bold', 12)
    p.drawString(x_start, start_y, f"Agendado")
    start_y -= 20
    # Desenha a lista de serviços do primeiro laço for
    end_y, end_page = draw_service_list(p, dados_servicos.filter(
        status='Agendado',
        data_inicio__gte=seven_days,
    ), start_y, start_page)

    # Adiciona uma nova página antes de iniciar o segundo laço for
    p.showPage()
    draw_header(p)
    start_y = height - 150  # Posição inicial para o conteúdo após o cabeçalho


    p.setFont('Helvetica-Bold', 12)
    p.drawString(x_start, start_y, f"Próximo")
    start_y -= 20
    # Exemplo do segundo laço for (dados_servicos_2)
    # Ajuste conforme necessário para seu segundo conjunto de dados
    end_y, end_page = draw_service_list(p, dados_servicos.filter(
        data_inicio__gte=one_day,
        data_inicio__lte=seven_days
    ).exclude(
        status="Em andamento",
        data_inicio__lte=seven_days
    ), start_y, end_page + 1)

    # Adiciona uma nova página antes de iniciar o segundo laço for
    p.showPage()
    draw_header(p)
    start_y = height - 150  # Posição inicial para o conteúdo após o cabeçalho

    p.setFont('Helvetica-Bold', 12)
    p.drawString(x_start, start_y, f"Volume de servicos agendados")
    start_y -= 20

    start_y, end_page = add_figures_to_pdf(p, fig_charts, start_y, end_page + 1)

    # Adiciona uma nova página antes de iniciar o segundo laço for
    p.showPage()
    draw_header(p)
    start_y = height - 150  # Posição inicial para o conteúdo após o cabeçalho

    p.setFont('Helvetica-Bold', 12)
    p.drawString(x_start, start_y, f"Volume de servicos próximos")
    start_y -= 20

    start_y, end_page = add_figures_to_pdf(p, fig_charts_proximos, start_y, end_page + 1)

    # Adiciona uma nova página antes de iniciar o segundo laço for
    p.showPage()
    draw_header(p)
    start_y = height - 150  # Posição inicial para o conteúdo após o cabeçalho

    p.setFont('Helvetica-Bold', 12)
    p.drawString(x_start, start_y, f"Volume de servicos em andamento")
    start_y -= 20

    start_y, end_page = add_figures_to_pdf(p, fig_charts_endamento, start_y, end_page + 1)

    # Adiciona uma nova página antes de iniciar o segundo laço for
    p.showPage()
    draw_header(p)
    start_y = height - 150  # Posição inicial para o conteúdo após o cabeçalho

    p.setFont('Helvetica-Bold', 12)
    p.drawString(x_start, start_y, f"Volume de servicos agendados por mês")
    start_y -= 20

    start_y, end_page = add_figures_to_pdf(p, fig_grouped_charts, start_y, end_page + 1)

    # Adiciona uma nova página antes de iniciar o segundo laço for
    p.showPage()
    draw_header(p)
    start_y = height - 150  # Posição inicial para o conteúdo após o cabeçalho

    # Desenha o rodapé na última página
    draw_footer(p, end_page, is_last_page=True)

    # Close the PDF object cleanly, and we're done.
    p.showPage()
    p.save()

    # Get the value of the BytesIO buffer and write it to the response.
    buffer.seek(0)

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio de planejamento de servicos.pdf"'

    return response


def exportar_excel_planejado(request, login_type, id, datastart, datafim, empresa, unidade, localidades, negocios):
    # Função que bloqueia a view em caso das variáveis de sessão terem sido alteradas no logout
    block = block_view(
        request,
        login_type=login_type,
        id=id
    )

    if block == True:
        return redirect('logout')

    # Crie um novo arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # cabeçalhos da tabela exportada
    headers = [
        'Tipo de empresa',
        'Empresa',

        # cabeçalhos referentes a tabela de serviços
        'ID Agendamento',
        'Tipo de agendamento',
        'Descrição do Serviço',
        'Colaboradores Chamados',
        'Serviços Solicitados',
        'Data de Início',
        # 'Data de conclusao',
        # 'Antes',
        # 'Depois',
        'Status do Serviço',

        # cabeçalhos referentes a tabela de equipamentos

        # 'Marca do Equipamento',
        # 'Equipamento catálogo',
        # 'Equipamento Empresa',
        # 'Tipo equipamento',
        # 'Equipamento id',
        # 'vida util equipamento (meses)',
        # 'Data de aquisição equipamento',
        # 'Data de desmobilização equipamento',
        # 'Matrícula do Equipamento',

        # cabeçalhos referentes a tabela de ferramentas

        # 'Marca da Ferramenta',
        # 'Ferramenta catálogo',
        # 'Ferramenta empresa',
        # 'Tipo ferramenta',
        # 'Ferramenta id',
        # 'vida util ferramenta (meses)',
        # 'Data de aquisição ferramenta',
        # 'Data de desmobilização ferramenta',
        # 'Matrícula da Ferramenta',

        # cabeçalhos referentes a tabela de materiais

        # 'Material Aplicado',
        # 'Material Categoria',
        # 'Forma de consumo',
        # 'Quantidade',
        # 'Origem do material',

        # cabeçalhos referentes a tabela de jardins
        'Área atendida',
        'Periodicidade',
        'Área total',
        'Tipo vegetação',
        'Tipo Terreno',
        'Localidade',
        'Negocio',
        'Unidade',

        # cabeçalhos referentes a tabela fato serviços
        # 'ID de serviço',
        # 'Tempo na área',
        # 'Colaborador envolvido',
        # 'Principal servico',
        # 'Data e hora de chagada',
        # 'Data e hora de retorno'
    ]

    # adiciona os cabeçahos no arquivo
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_title

    # Adicione os dados do relatório ao arquivo Excel
    dados = colect_dados_planejamento(

    )

    # aplica os filtros enviados como parametro na media que existirem
    if datastart != 'None':
        datastart = datetime.strptime(datastart, '%Y-%m-%d')

    if datafim != 'None':
        datafim = datetime.strptime(datafim, '%Y-%m-%d')

    if empresa != 'None':
        dados = dados.filter(
            empresaprestadora=empresa
        )
    else:
        pass

    if unidade != 'None':
        dados = dados.filter(
            unidade=unidade
        )
    else:
        pass

    if localidades != 'None':
        dados = dados.filter(
            localidade=localidades
        )
    else:
        pass

    if negocios != 'None':
        dados = dados.filter(
            tipodeempresa=negocios
        )
    else:
        pass

    if datastart != 'None':
        dados = dados.filter(
            data_de_inicio__gte=datastart  # Usar __gte para maior ou igual
        )
    else:
        pass


    if datafim != 'None':  # Mudar para datafim ao invés de datastart
        dados = dados.filter(
            data_de_inicio__lte=datafim  # Usar __lte para menor ou igual
        )
    else:
        pass


    # escreve as linhas do arquivo excel antes de ser exportado
    for row_num, row in enumerate(dados, start=2):
        row_data = [
            row.tipodeempresa, row.empresaprestadora,

            row.id_agendamento, row.tipo_agendamento,

            row.descricao_do_servico, row.colaboradores_chamados, row.servicos_solicitados, row.data_de_inicio,

            # row.data_de_conclusao, row.antes, row.depois,

            row.status_servico,

            # row.equipamento_marca, row.equipamento_catalogo, row.equipamento_empresa, row.tipo_equipamento, row.equipamento_id, row.vida_util_equipamento, row.data_aquisicao_equipamento, row.data_desmobilizacao_equipamento, row.matricula_equipamento,
            #
            # row.ferramenta_marca, row.ferramenta_catalogo, row.ferramenta_empresa, row.tipo_ferramenta, row.ferramenta_id, row.vida_util_ferramenta, row.data_aquisicao_ferramenta, row.data_desmobilizacao_ferramenta, row.matricula_ferramenta,

            # row.material_aplicado, row.material_categoria, row.forma_consumo, row.qtd, row.tipo_material,

            row.area_atendida, row.periodicidade_de_retorno,

            row.area_total, row.tipo_vegetacao, row.tipo_terreno, row.localidade, row.tiponegocio, row.unidade,

            # row.id_servico, row.tempo_na_area, row.colaborador_envolvido, row.principalservico,

            # row.data_hora_chegada, row.data_hora_retorno
        ]
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(
                row=row_num,
                column=col_num
            )
            cell.value = value

    # Defina o nome do arquivo e o tipo de resposta
    # cria uma resposta http com o tipo de conteuso
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # cria um anexo que ´instalado no lado cliente
    response['Content-Disposition'] = 'attachment; filename="relatorio de servicos planejados.xlsx"'

    # Salve o arquivo Excel
    wb.save(response)

    # retorna a resposta
    # linha obrigatória
    # ou o sistema quebra
    return response


def relatoriodeservicosplanejados(request):
    login_type = request.session['login_type']
    id = request.session['login_id']


    # Função que bloqueia a view em caso das variáveis de sessão terem sido alteradas no logout
    block = block_view(
        request,
        login_type=login_type,
        id=id
    )

    if block == True:
        return redirect('logout')


    empresas = None
    unidades = None
    localidades = None
    negocios = None
    datastart = None
    datafim = None

    # Adicione os dados do relatório ao arquivo Excel
    dados = colect_dados_planejamento().filter(
        status_servico__in = ["Agendado", "Em andamento"]
    )[:50]

    # formulário que envia os filtros para serem aplicados
    forms_filtro = FiltroTableForms()

    if request.method == 'POST':
        form = FiltroTableForms(request.POST)
        if form.is_valid():
            # Processar os dados do formulário
            # os dados do formulário não são enviados para o banco de dados
            # são processado como paramtros para serem aplicados como filtro
            empresas = form.cleaned_data['empresa']
            unidades = form.cleaned_data['unidade']
            localidades = form.cleaned_data.get('localidade')
            datastart = form.cleaned_data.get('datastart')
            datafim = form.cleaned_data.get('datafim')

            dados = colect_dados_planejamento().filter(
                status_servico__in=["Agendado", "Em andamento"]
            )
            # aplicação dos filtros na medida que for nescessário
            if empresas:
                dados = dados.filter(
                    empresaprestadora=str(empresas).split("|")[0].strip()
                )
                empresas = str(empresas).split("|")[0].strip()

            if unidades:
                dados = dados.filter(
                    unidade=unidades
                )
                unidades = unidades

            if localidades:
                dados = dados.filter(
                    localidade=str(localidades).split("|")[1].strip(),
                )
                localidades = str(localidades).split("|")[1].strip()


            if negocios:
                dados = dados.filter(
                    localidade=str(localidades).split("|")[2].strip(),
                )
                negocios = str(localidades).split("|")[2].strip()


            if datastart:
                dados = dados.filter(
                    data_de_inicio__gte=datastart  # Usar __gte para maior ou igual
                )

            if datafim:  # Mudar para datafim ao invés de datastart
                dados = dados.filter(
                    data_de_inicio__lte=datafim  # Usar __lte para menor ou igual
                )

    elementos_paginados = paginate(
        request=request,
        data_objects=dados,
        per_page=20
    )

    return render(request, 'controle/dashboards/relatorio_de_servicos_concluidos.html',
                  {'elementos_paginados': elementos_paginados,
                   'forms_filtro': forms_filtro,
                   'empresas': empresas,
                   'unidades': unidades,
                   'localidades': localidades,
                   'negocios':negocios,
                   'datastart': datastart,
                   'datafim': datafim
                   }
    )